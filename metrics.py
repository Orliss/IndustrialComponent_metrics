import os
import cv2
import json
import numpy as np
import random
from PIL import Image, ImageEnhance, ImageFilter
import random
from openpyxl import Workbook,load_workbook
import matplotlib.pyplot as plt
import torch
import torch.nn as nn

# 1. 编写模型效果评价脚本，人工标注的图像和标注会放在同一个文件夹下，如果是无裂纹图像就不会有标注文件，
# 有裂纹图像则会有一个同名的json文件。算法也会对这些图像全部推理，生成预测的json文件。
# 你的脚本要根据人工标注的GT json和模型预测的pred json，计算误检率和漏检率指标。
# 漏检定义为只要pred的裂纹与gt裂纹检测出来一点就不算漏检，
# 如果原来没有裂纹的地方检测出来一道pred裂纹就算误检。（这个要求本周六前完成）
def model_predict(pred_model,weight:str,file_path:str,save_path:str):
    '''
    Not_Finish
    '''
    LABELME_VERSION="2.3.6"
    imgsize=512

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    model=pred_model
    model.load_state_dict(torch.load(weight, map_location=device))
    model.eval()
    model.to(device)
    shapes=[]

    for file_name in os.listdir(file_path):
        if file_name.split('.')[-1] == 'jpg':
            img=cv2.imread(os.path.join(file_path, file_name),cv2.COLOR_BGR2RGB)
            ori_h, ori_w = img.shape[:2]
            img=cv2.resize(img,(imgsize,imgsize))
            # img = (img / 255.0 - mean) / std
            save_file=os.path.join(file_path, file_name.replace('.jpg', '_pred.json'))
            with torch.no_grad():
                img = torch.from_numpy(img).permute(2, 0, 1).float()  # HWC → CHW
                img = img.unsqueeze(0).to(device)

                pred_result=model(img)


        shapes.append({
            "group_id": None,
            "description": "",
            "difficult": False,
            "shape_type": "polygon",
            "flags": {},
            "attributes": {}
        })
        labelme_data = {
            "version": LABELME_VERSION,
            "flags": {},
            "shapes": shapes,
            "imagePath": file_name,
            "imageWidth": img.shape[1],
            "imageHeight": img.shape[0],
        }
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(labelme_data, f, ensure_ascii=False, indent=2)

def result_show(xlsx_path,
                pred_image_path,
                fused_image_path,
                gt_image_path,
                save_path):
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        fig, axs = plt.subplots(2, 2)

        image_filename = ws.cell(row=row_num, column=1).value
        result = ws.cell(row=row_num, column=5).value
        if ws.cell(row=row_num, column=6).value is not None:
            result = result + '_' + ws.cell(row=row_num, column=6).value

        pred_file_path = os.path.join(pred_image_path,image_filename.split('.')[0] + '.png')

        image_filename = image_filename.removeprefix('check_NG_').removeprefix('check_OK_').split('.')[0] + '.png'

        fused_file_path = os.path.join(fused_image_path,image_filename.split('.')[0] + '.jpg')
        gt_file_path = os.path.join(gt_image_path,image_filename)
        save_file_path = os.path.join(save_path,result + '_' + image_filename)

        fused_image = Image.open(fused_file_path)  # .resize((image_width, image_height), Image.Resampling.LANCZOS)
        pred_image = Image.open(pred_file_path)
        gt_image = Image.open(gt_file_path) if os.path.exists(gt_image_path) else None

        axs[0, 0].axis('off')
        axs[0, 0].imshow(fused_image)
        axs[0, 0].set_title('fused_image')

        axs[0, 1].axis('off')
        axs[0, 1].imshow(pred_image, cmap='gray', vmin=0, vmax=255)
        axs[0, 1].set_title('pred_image')

        # axs[1, 0].imshow()
        axs[1, 0].axis('off')
        axs[1, 0].text(0.5, 0.5, result, ha="center", va="center", fontsize=14)

        axs[1, 1].axis('off')
        axs[1, 1].imshow(gt_image, cmap='gray', vmin=0, vmax=255
                         ) if gt_image is not None else axs[1, 1].text(0.5, 0.5, 'no_gt', ha="center", va="center",
                                                                       fontsize=14)
        axs[1, 1].set_title('gt_image')
        plt.savefig(save_file_path, dpi=200, bbox_inches='tight')
        plt.close('all')
        # plt.show()

    wb.close()

def data_write_test(file_path:str,datalist):
    wb = Workbook()
    ws = wb.active
    ws.append(['file_name','exist_GTlabel','GT_num','exist_Predlabel','Pred_num','result'])
    for row in datalist:
        ws.append(row)

    wb.save(file_path+'_result.xlsx')
    print('done')

def metrics_cal_test(file_path:str,threshold:float=0.75):
    TP,TN=0,0
    FP,FN=0,0
    miss,correct,wrong=0,0,0
    # data=[file_name,exist_GTlabel,GT_num,exist_Predlabel,Pred_num,result]
    datalist=[]
    for file_name in os.listdir(file_path):
        data = []
        if file_name.split('.')[-1] == 'jpg':
            data.append(file_name)
            img=cv2.imread(os.path.join(file_path,file_name))
            h,w=img.shape[:2]

            #ground_truth json label
            GT_json_name=file_name.replace('.jpg', '.json')
            GT_json_path=os.path.join(file_path, GT_json_name)

            # predict json label
            Pred_json_name=file_name.replace('.jpg', '_pred.json')
            Pred_json_path=os.path.join(file_path, Pred_json_name)

            if os.path.exists(GT_json_path):
                data.append('True')
                with open(GT_json_path, 'r', encoding='utf-8') as f:
                    GT_data = json.load(f)['shapes']
                    GT_points=[np.array(_['points']) for _ in GT_data]
                    GT_num=len(GT_data)
                    data.append(GT_num)

                if os.path.exists(Pred_json_path):
                    data.append('True')
                    with open(Pred_json_path, 'r', encoding='utf-8') as f:
                        Pred_data = json.load(f)['shapes']
                        Pred_points = [np.array(_['points']) for _ in Pred_data]
                        Pred_num=len(Pred_data)
                        data.append(Pred_num)
                    # if GT_data['shape']!=Pred_data['shape']:
                    #     data.append('miss or wrong')
                    #[crack_1_img,crack_2_img,...]
                    GT_png=[cv2.fillPoly(np.zeros((h, w), dtype=np.uint8),
                                            [_.astype(np.int32)], 256) for _ in GT_points]
                    Pred_png = [cv2.fillPoly(np.zeros((h, w), dtype=np.uint8),
                                            [_.astype(np.int32)], 256) for _ in Pred_points]

                    #calculate IOU matrix
                    result=[]
                    for gt_ in GT_png:
                        _=[]
                        for pred_ in Pred_png:
                            #calaulate IOU
                            intersection= cv2.bitwise_and(gt_, pred_)
                            union=cv2.bitwise_or(gt_, pred_)

                            intersection_area = np.count_nonzero(intersection)
                            union_area = np.count_nonzero(union)

                            _.append(1 if (intersection_area / union_area)>threshold else 0)
                        result.append(_)
                    is_correct=False
                    for _ in result:
                        if sum(_)==1:
                            is_correct = True
                        elif sum(_)==0:
                            miss+=1
                            is_correct = False
                            data.append('miss')
                            break
                        elif sum(_)>1:
                            wrong+=1
                            is_correct = False
                            data.append('wrong')
                            break
                    if is_correct:
                        correct += 1
                        data.append('correct')
                else:
                    miss += 1
                    data.append('False')
                    data.append(0)
                    data.append('miss')
            else:
                data.append('False')
                data.append(0)
                if os.path.exists(Pred_json_path):
                    data.append('True')
                    with open(Pred_json_path, 'r', encoding='utf-8') as f:
                        Pred_data = json.load(f)['shapes']
                        data.append(len(Pred_data))
                    if len(Pred_data)==0:
                        correct+=1
                        data.append('correct')
                    else:
                        wrong += 1
                        data.append('wrong')
                else:
                    correct+=1
                    data.append('False')
        datalist.append(data)

    print('correct: ',correct)
    print('wrong:   ',wrong)
    print('miss:    ',miss)
    data_write(file_path,datalist)

def data_write(file_path:str,datalist,ver):
    '''
    use this
    '''
    wb = Workbook()
    ws = wb.active
    ws.append(['pred_label_name','pred_num','exist_gt_label','gt_num','result'])
    for row in datalist:
        ws.append(row)

    wb.save(os.path.join(file_path,ver+'_result.xlsx'))
    print('done')

def metrics_cal(pred_path,gt_path:str,save_path,ver,threshold:float=0.75):
    '''
    use this
    '''
    TP,TN=0,0
    FP,FN=0,0
    miss,correct,wrong=0,0,0
    # data=['pred_label_name','pred_num','gt_label_name','gt_num','result']
    datalist=[]
    for file_name in os.listdir(pred_path):
        data = [file_name]
        Pred_json_path = os.path.join(pred_path, file_name)
        with open(Pred_json_path,'r',encoding='utf-8') as f:
            Pred_data=json.load(f)
            Pred_points = [np.array(_['points']) for _ in Pred_data['shapes']]
            Pred_num = len(Pred_data['shapes'])
            data.append(Pred_num)

            img_name=Pred_data["imagePath"]
            h=Pred_data["imageHeight"]
            w=Pred_data["imageWidth"]

        if os.path.exists(os.path.join(gt_path,
                                       file_name.split('.')[0].removeprefix('check_NG_').removeprefix('check_OK_')+'.json')):
            # print(file_name)
            data.append('True')
            with open(os.path.join(gt_path,
                                       file_name.split('.')[0].removeprefix('check_NG_').removeprefix('check_OK_')+'.json'),
                      'r', encoding='utf-8') as f:
                GT_data = json.load(f)['shapes']
                GT_points = [np.array(_['points']) for _ in GT_data]
                GT_num = len(GT_data)
                data.append(GT_num)

                if Pred_num == 0:
                    miss += 1
                    data.append('miss')
                    datalist.append(data)
                    continue

                #[crack_1_img,crack_2_img,...]
                GT_png=[cv2.fillPoly(np.zeros((h, w), dtype=np.uint8),
                                        [_.astype(np.int32)], 256) for _ in GT_points]
                Pred_png = [cv2.fillPoly(np.zeros((h, w), dtype=np.uint8),
                                        [_.astype(np.int32)], 256) for _ in Pred_points]

                #calculate IOU matrix
                result=[]
                for pred_ in Pred_png:
                    _=[]
                    for gt_ in GT_png:
                        #calaulate IOU
                        intersection= cv2.bitwise_and(gt_, pred_)
                        union=cv2.bitwise_or(gt_, pred_)

                        intersection_area = np.count_nonzero(intersection)
                        union_area = np.count_nonzero(union)

                        _.append(1 if (intersection_area / union_area)>threshold else 0)
                    result.append(_)

                is_correct_wrong=is_correct_miss=False

                for _ in result:
                    if sum(_)==1:
                        is_correct_wrong = True
                    elif sum(_)==0:
                        wrong+=1
                        is_correct_wrong = False
                        data.append('wrong')
                        break

                for _T in np.array(result).T:
                    if _T.sum()>0:
                        is_correct_miss = True
                    else:
                        miss += 1
                        is_correct_miss = False
                        data.append('miss')
                        break

                if is_correct_wrong and is_correct_miss:
                    correct += 1
                    data.append('correct')
                    # print(file_name)
            # else:
            #     miss += 1
            #     data.append('False')
            #     data.append(0)
            #     data.append('miss')
        else:
            data.append('False')
            data.append('0')
            if Pred_num==0:
                correct+=1
                data.append('correct')
                # print(file_name)
            else:
                wrong += 1
                data.append('wrong')

        datalist.append(data)

    print('correct: ',correct)
    print('wrong:   ',wrong)
    print('miss:    ',miss)
    data_write(save_path,datalist,ver)

# ---------------------------------------------------------------------------------

# 2. 寻找一些工业检测数据集中的裂纹标注，
# 提取这些裂纹的形状掩码，在我们采集的无伤工件上人为构造裂纹缺陷的样本，相当于做数据增强的工作。
# 这个分两步来，第一步先找合适的数据集，第二步写脚本实现无伤样本数据增强。你评估一下需要多长时间。
def template_generate():
    '''
    Not_Finish
    '''
    pass

def min_outline_cal(img_path):
    '''
    Not_Finish
    '''

    img=cv2.imread(img_path,cv2.IMREAD_UNCHANGED)
    gray_img=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    ret, binary = cv2.threshold(gray_img, 64, 255,
                                cv2.THRESH_BINARY_INV)
    contours,hierarchy=cv2.findContours(binary,
                                        cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    min_rect = cv2.minAreaRect(max(contours, key=cv2.contourArea))
    box=np.intp(cv2.boxPoints(min_rect))
    cv2.drawContours(img, [box], 0, (0, 0, 255), 2)
    cv2.imshow('binary',img)
    cv2.waitKey(0)


    return img


def template_enhance(img_path,enhancer,threshold):
    img=cv2.imread(img_path,cv2.IMREAD_COLOR_RGB)
    for e,t in zip(enhancer,threshold):
        if e in [ImageEnhance.Brightness,ImageEnhance.Contrast,ImageEnhance.Color]:
            img = Image.fromarray(img.astype(np.uint8))
            img=np.array(e(img).enhance(t))
        elif e in ['flip_lr','flip_ud','rotate','noise']:
            img=Image.fromarray(img)
            if e == 'flip_lr':
                img=img.transpose(Image.Transpose.FLIP_LEFT_RIGHT)
            elif e == 'flip_ud':
                img=img.transpose(Image.Transpose.FLIP_TOP_BOTTOM)
            elif e == 'rotate':
                img=img.rotate(random.randint(*t), expand=True)
            elif e == 'noise':
                img_array = np.array(img)
                noise = np.random.normal(t[0],t[1], img_array.shape).astype(np.uint8)
                img=Image.fromarray(np.clip(img_array + noise, 0, 255).astype(np.uint8))
    img=np.array(img)
    #calculate min_rectangle box
    img=min_outline_cal(img)
    return img

def img_augment(file_path,template_path,
                template_enhancer,template_threshold):
    '''
    Not_Finish
    '''
    # template_enhancer=[ImageEnhance.Brightness,ImageEnhance.Contrast,'rotare',ImageEnhance.Color,'noise']
    # template_threshold=[0.5,1.8,(-90,30),2.0,[0,20]]
    template_img = []
    for file_name in os.listdir(template_path):
        template_img.append(os.path.join(template_path,file_name))
    for file_name in os.listdir(file_path):
        if file_name.split('.')[-1] == 'jpg':
            raw_img=cv2.imread(os.path.join(file_path, file_name))
            tp_png=template_enhance(random.choice(template_img),template_enhancer,template_threshold)
            h,w=tp_png.shape[:2]
            x, y = (random.randint(w, raw_img.shape[1]-w),
                    random.randint(h, raw_img.shape[0]-h))
            raw_img[y:y + h, x:x + w] = tp_png
            # cv2.namedWindow('test', cv2.WINDOW_NORMAL)
            # cv2.resizeWindow('test', 800, 600)
            # cv2.imshow('test',tp_png)
            # cv2.waitKey(0)




if __name__ == '__main__':
    # random.seed(114514)
    # np.random.seed(114514)

    metrics_cal(r'pred_path',
                r'gt_path',
                r'save_path',
                'ver',0)

    # min_outline_cal(r'.\111.png')
    # img_augment(r'\no_crack_img',
    #             r'\template_img'
    #             ,None,
    #             None)








