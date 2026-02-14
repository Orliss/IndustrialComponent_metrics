import os
import cv2
import json
import numpy as np
import random
from PIL import Image, ImageEnhance, ImageFilter
import random
from openpyxl import Workbook,load_workbook
import matplotlib.pyplot as plt
from matplotlib.collections import LineCollection,PatchCollection
from scipy.interpolate import make_interp_spline
from matplotlib.patches import Polygon
import torch
import torch.nn as nn

# 1. 编写模型效果评价脚本，人工标注的图像和标注会放在同一个文件夹下，如果是无裂纹图像就不会有标注文件，
# 有裂纹图像则会有一个同名的json文件。算法也会对这些图像全部推理，生成预测的json文件。
# 你的脚本要根据人工标注的GT json和模型预测的pred json，计算误检率和漏检率指标。
# 漏检定义为只要pred的裂纹与gt裂纹检测出来一点就不算漏检，
# 如果原来没有裂纹的地方检测出来一道pred裂纹就算误检。（这个要求本周六前完成）
def model_predict(pred_model,weight:str,file_path:str,save_path:str):
    '''
    Not_Finish
    '''
    LABELME_VERSION="2.3.6"
    imgsize=512

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    model=pred_model
    model.load_state_dict(torch.load(weight, map_location=device))
    model.eval()
    model.to(device)
    shapes=[]

    for file_name in os.listdir(file_path):
        if file_name.split('.')[-1] == 'jpg':
            img=cv2.imread(os.path.join(file_path, file_name),cv2.COLOR_BGR2RGB)
            ori_h, ori_w = img.shape[:2]
            img=cv2.resize(img,(imgsize,imgsize))
            # img = (img / 255.0 - mean) / std
            save_file=os.path.join(file_path, file_name.replace('.jpg', '_pred.json'))
            with torch.no_grad():
                img = torch.from_numpy(img).permute(2, 0, 1).float()  # HWC → CHW
                img = img.unsqueeze(0).to(device)

                pred_result=model(img)


        shapes.append({
            "group_id": None,
            "description": "",
            "difficult": False,
            "shape_type": "polygon",
            "flags": {},
            "attributes": {}
        })
        labelme_data = {
            "version": LABELME_VERSION,
            "flags": {},
            "shapes": shapes,
            "imagePath": file_name,
            "imageWidth": img.shape[1],
            "imageHeight": img.shape[0],
        }
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(labelme_data, f, ensure_ascii=False, indent=2)

def result_show(xlsx_path,
                pred_image_path,
                fused_image_path,
                gt_image_path,
                save_path):
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        fig, axs = plt.subplots(2, 2)

        image_filename = ws.cell(row=row_num, column=1).value
        result = ws.cell(row=row_num, column=5).value
        if ws.cell(row=row_num, column=6).value is not None:
            result = result + '_' + ws.cell(row=row_num, column=6).value

        pred_file_path = os.path.join(pred_image_path,image_filename.split('.')[0] + '.png')

        image_filename = image_filename.removeprefix('check_NG_').removeprefix('check_OK_').split('.')[0] + '.png'

        fused_file_path = os.path.join(fused_image_path,image_filename.split('.')[0] + '.jpg')
        gt_file_path = os.path.join(gt_image_path,image_filename)
        save_file_path = os.path.join(save_path,result + '_' + image_filename)

        fused_image = Image.open(fused_file_path)  # .resize((image_width, image_height), Image.Resampling.LANCZOS)
        pred_image = Image.open(pred_file_path)
        gt_image = Image.open(gt_file_path) if os.path.exists(gt_image_path) else None

        axs[0, 0].axis('off')
        axs[0, 0].imshow(fused_image)
        axs[0, 0].set_title('fused_image')

        axs[0, 1].axis('off')
        axs[0, 1].imshow(pred_image, cmap='gray', vmin=0, vmax=255)
        axs[0, 1].set_title('pred_image')

        # axs[1, 0].imshow()
        axs[1, 0].axis('off')
        axs[1, 0].text(0.5, 0.5, result, ha="center", va="center", fontsize=14)

        axs[1, 1].axis('off')
        axs[1, 1].imshow(gt_image, cmap='gray', vmin=0, vmax=255
                         ) if gt_image is not None else axs[1, 1].text(0.5, 0.5, 'no_gt', ha="center", va="center",
                                                                       fontsize=14)
        axs[1, 1].set_title('gt_image')
        plt.savefig(save_file_path, dpi=200, bbox_inches='tight')
        plt.close('all')
        # plt.show()

    wb.close()

def data_write(file_path:str,datalist,ver):
    '''
    use this
    '''
    wb = Workbook()
    ws = wb.active
    ws.append(['pred_label_name','pred_num','exist_gt_label','gt_num','result'])
    for row in datalist:
        ws.append(row)

    wb.save(os.path.join(file_path,ver+'_result.xlsx'))
    print('done')

def metrics_cal(pred_path,gt_path:str,save_path,ver,threshold:float=0.75):
    '''
    use this
    '''
    TP,TN=0,0
    FP,FN=0,0
    miss,correct,wrong=0,0,0
    # data=['pred_label_name','pred_num','gt_label_name','gt_num','result']
    datalist=[]
    for file_name in os.listdir(pred_path):
        data = [file_name]
        Pred_json_path = os.path.join(pred_path, file_name)
        with open(Pred_json_path,'r',encoding='utf-8') as f:
            Pred_data=json.load(f)
            Pred_points = [np.array(_['points']) for _ in Pred_data['shapes']]
            Pred_num = len(Pred_data['shapes'])
            data.append(Pred_num)

            img_name=Pred_data["imagePath"]
            h=Pred_data["imageHeight"]
            w=Pred_data["imageWidth"]

        if os.path.exists(os.path.join(gt_path,
                                       file_name.split('.')[0].removeprefix('check_NG_').removeprefix('check_OK_')+'.json')):
            # print(file_name)
            data.append('True')
            with open(os.path.join(gt_path,
                                       file_name.split('.')[0].removeprefix('check_NG_').removeprefix('check_OK_')+'.json'),
                      'r', encoding='utf-8') as f:
                GT_data = json.load(f)['shapes']
                GT_points = [np.array(_['points']) for _ in GT_data]
                GT_num = len(GT_data)
                data.append(GT_num)

                if Pred_num == 0:
                    miss += 1
                    data.append('miss')
                    datalist.append(data)
                    continue

                #[crack_1_img,crack_2_img,...]
                GT_png=[cv2.fillPoly(np.zeros((h, w), dtype=np.uint8),
                                        [_.astype(np.int32)], 256) for _ in GT_points]
                Pred_png = [cv2.fillPoly(np.zeros((h, w), dtype=np.uint8),
                                        [_.astype(np.int32)], 256) for _ in Pred_points]

                #calculate IOU matrix
                result=[]
                for pred_ in Pred_png:
                    _=[]
                    for gt_ in GT_png:
                        #calaulate IOU
                        intersection= cv2.bitwise_and(gt_, pred_)
                        union=cv2.bitwise_or(gt_, pred_)

                        intersection_area = np.count_nonzero(intersection)
                        union_area = np.count_nonzero(union)

                        _.append(1 if (intersection_area / union_area)>threshold else 0)
                    result.append(_)

                is_correct_wrong=is_correct_miss=False

                for _ in result:
                    if sum(_)==1:
                        is_correct_wrong = True
                    elif sum(_)==0:
                        wrong+=1
                        is_correct_wrong = False
                        data.append('wrong')
                        break

                for _T in np.array(result).T:
                    if _T.sum()>0:
                        is_correct_miss = True
                    else:
                        miss += 1
                        is_correct_miss = False
                        data.append('miss')
                        break

                if is_correct_wrong and is_correct_miss:
                    correct += 1
                    data.append('correct')
                    # print(file_name)
            # else:
            #     miss += 1
            #     data.append('False')
            #     data.append(0)
            #     data.append('miss')
        else:
            data.append('False')
            data.append('0')
            if Pred_num==0:
                correct+=1
                data.append('correct')
                # print(file_name)
            else:
                wrong += 1
                data.append('wrong')

        datalist.append(data)

    print('correct: ',correct)
    print('wrong:   ',wrong)
    print('miss:    ',miss)
    data_write(save_path,datalist,ver)

# ---------------------------------------------------------------------------------

# 2. 寻找一些工业检测数据集中的裂纹标注，
# 提取这些裂纹的形状掩码，在我们采集的无伤工件上人为构造裂纹缺陷的样本，相当于做数据增强的工作。
# 这个分两步来，第一步先找合适的数据集，第二步写脚本实现无伤样本数据增强。你评估一下需要多长时间。

class LinearTemplateGenerate_v1:
    def __init__(self, start_pos, start_angle,
                 length, base_width,
                 depth=0,max_depth=5,
                 branch_mode='linear'
                 ):
        self.points = [start_pos]
        self.angle = start_angle
        self.length = length
        self.base_width = base_width
        self.children = []
        self.depth=depth
        self.max_depth=max_depth

        self.trend = random.choice([-1, 1])
        self.line_mode = random.choice(['A', 'B', 'C'])
        self.branch_mode = branch_mode

    def single_gen(self):
        current_pos = self.points[0]
        current_angle = self.angle

        num_steps = 15
        step_size = self.length / num_steps

        for i in range(num_steps):
            angle_noise = np.random.uniform(-10, 10)
            current_angle += angle_noise
            radian = np.radians(current_angle)

            dx = step_size * np.cos(radian)
            dy = step_size * np.sin(radian)

            next_pos = (current_pos[0] + dx, current_pos[1] + dy)
            self.points.append(next_pos)
            current_pos = next_pos


    def multi_gen(self):
        current_pos = self.points[0]
        current_angle = self.angle

        num_steps = 10
        step_size = self.length / num_steps

        for i in range(num_steps):
            angle_noise = np.random.uniform(-10, 10)
            current_angle += angle_noise
            radian=np.radians(current_angle)

            dx = step_size * np.cos(radian)
            dy = step_size * np.sin(radian)

            next_pos = (current_pos[0] + dx, current_pos[1] + dy)
            self.points.append(next_pos)
            current_pos = next_pos

            if (# 0.3 < i / num_steps < 0.7 and
                self.depth < self.max_depth and
                np.random.random() < 0.15):

                current_tval = (i + 1) / num_steps
                factor = self.width_factor_cal(current_tval)
                connect_width = self.base_width * factor
                current_branch_mode=random.choices(['linear','random'],
                                            [8,2],k=1)[0]if self.branch_mode=='mix'else self.branch_mode
                if current_branch_mode=='linear':
                    offset_angle = current_angle+ np.random.uniform(-45, 45)
                elif current_branch_mode== 'random':
                    offset_angle = current_angle+ np.random.uniform(-180, 180)
                # branch_angle=np.random.choice([-1,1])*np.radians(offset_degree)

                child = LinearTemplateGenerate_v1(current_pos, offset_angle,
                                    random.uniform(self.length * 0.6,self.length * 1.4),
                                    connect_width,
                                    # random.uniform(self.base_width * 0.6, self.base_width * 1.2),
                                    depth=self.depth + 1,
                                    max_depth=self.max_depth-1,
                                    branch_mode=self.branch_mode)
                child.multi_gen()
                self.children.append(child)

    def width_factor_cal(self, t_val):
        if self.line_mode == 'A':
            return np.sin(t_val * np.pi)
        elif self.line_mode == 'B':
            return (1 - t_val) ** 0.5
        elif self.line_mode == 'C':
            taper = (1 + self.trend * t_val * 0.6)
            bulge = 1 + 0.3 * np.sin(t_val * 15)
            shape_factor = taper * bulge
            return shape_factor
        else :
            return 1
            # start_adjust = (1 + self.trend * 0) * (1 + 0.2 * np.sin(0))
            # return (taper * bulge) / start_adjust

    def render(self):

        all_lines = []
        all_widths = []


        pts = np.array(self.points)
        x = pts[:, 0]
        y = pts[:, 1]

        if len(x) < 3:
            return [], []

        t = np.linspace(0, 1, len(x))
        t_smooth = np.linspace(0, 1, 200)  # 插值成 200 个细微点

        try:
            spl_x = make_interp_spline(t, x, k=2)(t_smooth)
            spl_y = make_interp_spline(t, y, k=2)(t_smooth)
        except:
            return [], []

        points_smooth = np.column_stack([spl_x, spl_y])
        segments = np.array([points_smooth[:-1], points_smooth[1:]]).transpose(1, 0, 2)
        t_vals = np.linspace(0, 1,len(segments))

        widths = []
        for t_val in t_vals:
            shape_factor=self.width_factor_cal(t_val)
            current_w = self.base_width * shape_factor
            widths.append(current_w)

        all_lines.extend(segments)
        all_widths.extend(widths)

        for child in self.children:
            c_lines, c_widths = child.render()
            all_lines.extend(c_lines)
            all_widths.extend(c_widths)

        return all_lines, all_widths

class TemplateGenerate:
    def __init__(self):
        # General
        self.points = []
        self.children = []
        self.style = "default"

        self.angle = 0
        self.length = 10
        self.base_width = 10
        self.depth=0
        self.max_depth=1
        self.trend = random.choice([-1, 1])
        self.line_style = random.choice(['spindle', 'tadpole', 'wave'])
        self.branch_style = 'default'
        self.strict_angle = 10

        self.center = (0.5,0.5)
        self.radius = 2
        # self.base_width = 1
        self.ring_num = 1
        self.distort = 0.1
        self.fill = True
        self.contours = []


    @classmethod
    def tree(cls, start_pos=(0,0),start_angle=45,
             length=1, base_width=1,
             depth=1,max_depth=5,strict_angle=30,
             branch_style='strict'):
        obj = cls()
        obj.points = [start_pos]
        obj.children = []
        obj.style = 'tree'

        obj.angle = start_angle
        obj.length = length
        obj.base_width = base_width
        obj.depth=depth
        obj.max_depth=max_depth
        obj.trend = random.choice([-1, 1])
        obj.line_style = random.choice(['spindle', 'tadpole', 'wave'])
        obj.branch_style = branch_style
        obj.strict_angle = strict_angle
        return obj

    @classmethod
    def pit(cls, center, radius, base_width, ring_num=1, distort=0.1,fill:bool=True):
        obj = cls()
        obj.style = 'pit'
        obj.center = center
        obj.radius = radius
        obj.base_width = base_width
        obj.ring_num = ring_num if not fill else 1
        obj.distort = distort
        obj.fill = fill
        obj.contours = []

        if fill:
            obj.ring_num = 100
        else:
            obj.ring_num = ring_num

        return obj

    def generate(self):
        if self.style =='tree':
            self._gen_tree()
            lines, widths, colors = self.render()
            lc = LineCollection(lines, linewidths=widths,
                                color=colors, alpha=1, capstyle='round')
            return lc

        elif self.style == 'pit':
            self._gen_pit()
            patches = self.render()  # 获取 Polygon 对象列表
            pc = PatchCollection(patches, match_original=False)

            if self.fill:
                # [修改 3] 生成渐变色列表
                n = len(patches)
                color_list = []
                for i in range(n):

                    t = i / (n - 1) if n > 1 else 0.5

                    r = 1 - (1.0 - 0.0) * t  # R: 0 -> 1
                    g = 1.0 - (1.0 - 1.0) * t  # G: 保持 1
                    b = 1 - (1.0 - 0.0) * t  # B: 0 -> 1

                    color_list.append((r, g, b))

                # 设置填充颜色列表，去掉边框
                pc.set_facecolor(color_list)
                pc.set_edgecolor('none')
            else:
                # 轮廓模式：设置边框颜色，无面颜色
                pc.set_facecolor('none')
                pc.set_edgecolor('black')
                pc.set_linewidth(self.base_width)

            return pc

    def _width_factor_cal(self, t_val):
        if self.line_style == 'spindle':
            return np.sin(t_val * np.pi)
        elif self.line_style == 'tadpole':
            return (1 - t_val) ** 0.5
        elif self.line_style == 'wave':
            taper = (1 + self.trend * t_val * 0.6)
            bulge = 1 + 0.3 * np.sin(t_val * 15)
            shape_factor = taper * bulge
            return shape_factor
        else :
            return 1
            # start_adjust = (1 + self.trend * 0) * (1 + 0.2 * np.sin(0))
            # return (taper * bulge) / start_adjust

    def _gen_tree(self):
        current_pos = self.points[0]
        current_angle = self.angle

        num_steps = 15
        step_size = self.length / num_steps

        for i in range(num_steps):
            angle_noise = np.random.uniform(-10, 10)
            current_angle += angle_noise
            radian = np.radians(current_angle)

            dx = step_size * np.cos(radian)
            dy = step_size * np.sin(radian)

            next_pos = (current_pos[0] + dx, current_pos[1] + dy)
            self.points.append(next_pos)
            current_pos = next_pos

            if (#0.3<i/num_steps<0.7 and
                    self.max_depth>1 and
                    self.depth < self.max_depth  and
                    np.random.random() < 0.1):

                current_tval = (i + 1) / num_steps
                factor = self._width_factor_cal(current_tval)
                connect_width = self.base_width * factor

                current_branch_style=random.choices(['strict','random'],
                                            [8,2],k=1)[0]if self.branch_style=='mix'else self.branch_style
                if current_branch_style=='strict':
                    offset_angle = current_angle+ np.random.uniform(-self.strict_angle, self.strict_angle)
                elif current_branch_style== 'random':
                    offset_angle = current_angle+ np.random.uniform(-180, 180)
                else:
                    offset_angle=current_angle
                # branch_angle=np.random.choice([-1,1])*np.radians(offset_degree)



                child = TemplateGenerate.tree(
                    start_pos=current_pos,
                    start_angle=offset_angle,
                    length=self.length*random.uniform(0.8, 1.2),
                    base_width=connect_width,
                    depth=self.depth+1,
                    max_depth=self.max_depth,
                    strict_angle=self.strict_angle,
                    branch_style=self.branch_style
                      # 传入像素宽度
                )
                # child.depth = self.depth + 1
                # child.max_depth = self.max_depth-1
                child._gen_tree()
                self.children.append(child)

    def _gen_pit(self):
        center_x, center_y = self.center
        for ring_i in range(self.ring_num):
            progress = ring_i / self.ring_num if self.ring_num > 1 else 0

            # [修改 2] 调整半径衰减逻辑
            # 为了让渐变填满整个圆，我们需要半径从 100% 缩减到接近 0%
            if self.fill:
                # 填充模式：线性缩减到 0
                base_r = self.radius * (1 - progress)
            else:
                # 线框模式：只缩减一点点，产生涟漪感
                base_r = self.radius * (1 - progress * 0.6)

            # 防止半径变为负数或0（会导致插值报错）
            if base_r <= 0.1:
                base_r = 0.1

            # 生成噪声点
            angles = np.linspace(0, 2 * np.pi, 12, endpoint=False)
            noise = np.random.uniform(1 - self.distort, 1 + self.distort, 12)
            radii = base_r * noise

            # 闭环插值
            angles_pad = np.concatenate([angles[-3:] - 2 * np.pi, angles, angles[:3] + 2 * np.pi])
            radii_pad = np.concatenate([radii[-3:], radii, radii[:3]])

            fine_angles = np.linspace(0, 2 * np.pi, 100)
            spl = make_interp_spline(angles_pad, radii_pad, k=3)
            fine_radii = spl(fine_angles)

            x = center_x + fine_radii * np.cos(fine_angles)
            y = center_y + fine_radii * np.sin(fine_angles)

            pts = np.column_stack([x, y])
            self.contours.append(pts)

    def render(self):
        color=(0,0,255)
        if self.style =='tree':
            all_lines = []
            all_widths = []
            all_colors = []

            pts = np.array(self.points)

            if len(pts) < 3:
                return [], [],[]

            t = np.linspace(0, 1, len(pts))
            t_smooth = np.linspace(0, 1, 200)  # 插值成 200 个细微点

            try:
                spl_x = make_interp_spline(t, pts[:, 0], k=2)(t_smooth)
                spl_y = make_interp_spline(t, pts[:, 1], k=2)(t_smooth)
            except:
                return [], [], []

            points_smooth = np.column_stack([spl_x, spl_y])
            segments = np.array([points_smooth[:-1], points_smooth[1:]]).transpose(1, 0, 2)
            t_vals = np.linspace(0, 1, len(segments))

            widths = []
            for t_val in t_vals:
                shape_factor = self._width_factor_cal(t_val)
                current_w = self.base_width * shape_factor
                widths.append(current_w)
                r = 0 + 0.6* t_val
                g = 0.7 + 0.3 * t_val
                b = 0 + 0.6 * t_val
                current_color = (r, g, b)

                all_colors.append(current_color)
            all_lines.extend(segments)
            all_widths.extend(widths)

            for child in self.children:
                c_lines, c_widths,c_colors = child.render()
                all_lines.extend(c_lines)
                all_widths.extend(c_widths)
                all_colors.extend(c_colors)

            return all_lines, all_widths,all_colors
        elif self.style =='pit':
            # [修改] 返回 Polygon 对象列表
            patches = []
            for pts in self.contours:
                # 创建多边形，closed=True 闭合路径
                poly = Polygon(pts, closed=True)
                patches.append(poly)
            return patches

def template_enhance(img_path,enhancer,threshold):
    img=cv2.imread(img_path,cv2.IMREAD_COLOR_RGB)
    for e,t in zip(enhancer,threshold):
        if e in [ImageEnhance.Brightness,ImageEnhance.Contrast,ImageEnhance.Color]:
            img = Image.fromarray(img.astype(np.uint8))
            img=np.array(e(img).enhance(t))
        elif e in ['flip_lr','flip_ud','rotate','noise']:
            img=Image.fromarray(img)
            if e == 'flip_lr':
                img=img.transpose(Image.Transpose.FLIP_LEFT_RIGHT)
            elif e == 'flip_ud':
                img=img.transpose(Image.Transpose.FLIP_TOP_BOTTOM)
            elif e == 'rotate':
                img=img.rotate(random.randint(*t), expand=True)
            elif e == 'noise':
                img_array = np.array(img)
                noise = np.random.normal(t[0],t[1], img_array.shape).astype(np.uint8)
                img=Image.fromarray(np.clip(img_array + noise, 0, 255).astype(np.uint8))
    img=np.array(img)
    #calculate min_rectangle box
    img=min_outline_cal(img)
    return img

def test_plot():
    fig, ax = plt.subplots(figsize=(2,2), dpi=300)
    fig.patch.set_facecolor('none')
    ax.set_facecolor('none')
    ax.margins(0.1)
    ax.set_aspect('equal')
    ax.axis('off')
    plt.tight_layout()


    a=TemplateGenerate.tree(start_pos=(0,0),start_angle=30,length=2, base_width=2,
                            max_depth=2,branch_style='strict')
    lc=a.generate()

    b = TemplateGenerate.pit(center=(2, 2),  radius=1,base_width=1,
        distort=0.2,fill=False)
    pit_filled = b.generate()

    ax.add_collection(lc)
    # ax.add_collection(pit_filled)

    ax.autoscale()
    plt.savefig(r'../Test/test1.png',
                dpi=300,
                transparent=True,
                bbox_inches='tight',
                pad_inches=1,

                facecolor='none',
                edgecolor='none'
                )


    plt.show()


def min_outline_cal(img_path):
    '''
    Not_Finish
    '''

    img=cv2.imread(img_path,cv2.IMREAD_UNCHANGED)
    gray_img=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    ret, binary = cv2.threshold(gray_img, 64, 255,
                                cv2.THRESH_BINARY_INV)
    contours,hierarchy=cv2.findContours(binary,
                                        cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    min_rect = cv2.minAreaRect(max(contours, key=cv2.contourArea))
    box=np.intp(cv2.boxPoints(min_rect))
    cv2.drawContours(img, [box], 0, (0, 0, 255), 2)
    cv2.imshow('binary',img)
    cv2.waitKey(0)


    return img

def img_augment(file_path,template_path,
                template_enhancer,template_threshold):
    '''
    Not_Finish
    '''
    # template_enhancer=[ImageEnhance.Brightness,ImageEnhance.Contrast,'rotare',ImageEnhance.Color,'noise']
    # template_threshold=[0.5,1.8,(-90,30),2.0,[0,20]]
    template_img = []
    for file_name in os.listdir(template_path):
        template_img.append(os.path.join(template_path,file_name))
    for file_name in os.listdir(file_path):
        if file_name.split('.')[-1] == 'jpg':
            raw_img=cv2.imread(os.path.join(file_path, file_name))
            tp_png=template_enhance(random.choice(template_img),template_enhancer,template_threshold)
            h,w=tp_png.shape[:2]
            x, y = (random.randint(w, raw_img.shape[1]-w),
                    random.randint(h, raw_img.shape[0]-h))
            raw_img[y:y + h, x:x + w] = tp_png
            # cv2.namedWindow('test', cv2.WINDOW_NORMAL)
            # cv2.resizeWindow('test', 800, 600)
            # cv2.imshow('test',tp_png)
            # cv2.waitKey(0)




if __name__ == '__main__':
    # random.seed(114514)
    # np.random.seed(114514)

    test_plot()

    # metrics_cal(r'pred_path',
    #             r'gt_path',
    #             r'save_path',
    #             'ver',0)

    # min_outline_cal(r'.\111.png')
    # img_augment(r'\no_crack_img',
    #             r'\template_img'
    #             ,None,
    #             None)

    # print('0123总误检率： ',(514+19)/1246)
    # print('0123总漏检率： ',(11+19)/1246)
    # print('0123gt误检率： ',(18+19)/59)
    # print('0123gt漏检率： ',(11+19)/59)
    # print('0123无gt误检率： ',496/(1246-59))
    # print('0123无gt漏检率： ',0/(1246-59))
    # print('')
    # print('0126总误检率： ',(746+19)/1246)
    # print('0126总漏检率： ',(4+19)/1246)
    # print('0126gt误检率： ',(28+19)/59)
    # print('0126gt漏检率： ',(4+19)/59)
    # print('0126无gt误检率： ',736/(1246-59))
    # print('0126无gt漏检率： ',0/(1246-59))






